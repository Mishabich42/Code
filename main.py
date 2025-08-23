import requests
from bs4 import BeautifulSoup
import openpyxl
import time
import concurrent.futures


def get_item_prices(item_code, server="europe"):
    """
    Отримує ціни на предмет з сайту. Якщо для міста є кілька цін,
    повертає тільки найнижчу.
    """
    if "_LEVEL" in item_code and item_code[-1].isdigit():
        enchantment_level = item_code[-1]
        url_param = f"{item_code}@{enchantment_level}"
    else:
        url_param = item_code

    url = f"https://albiononlinetools.com/economy/item.php?itemPrice={url_param}&server={server}"
    price_data = {}

    try:
        response = requests.get(url, timeout=15)
        if response.status_code != 200:
            print(f"  -> Помилка запиту для {item_code}: Статус {response.status_code}")
            return {}

        soup = BeautifulSoup(response.text, "html.parser")
        table = soup.find(class_="table-responsive")

        if not table or not table.find("tbody"):
            print(f"  -> Не знайдено таблицю з цінами для {item_code}")
            return {}

        rows = table.find("tbody").find_all("tr")
        for row in rows:
            cols = [c.get_text(strip=True) for c in row.find_all("td")]
            if len(cols) > 1:
                city_name = cols[1].split(" (")[0]
                price_str = cols[2]

                try:
                    current_price = int(price_str.replace(',', ''))
                except ValueError:
                    continue

                if city_name not in price_data or current_price < price_data[city_name]:
                    price_data[city_name] = current_price

    except requests.RequestException as e:
        print(f"  -> Помилка мережі при обробці {item_code}: {e}")
        return {}

    return price_data


def update_sheet(sheet):
    """
    Оновлює ціни для одного конкретного аркуша (sheet).
    """
    print(f"\n--- Початок оновлення аркуша: '{sheet.title}' ---")
    header = [cell.value for cell in sheet[1]]
    city_columns = {city_name: col_idx + 1 for col_idx, city_name in enumerate(header) if
                    city_name and city_name != 'Item'}

    if not city_columns:
        print(f"-> Пропускаємо аркуш '{sheet.title}', оскільки він не має стовпців з назвами міст.")
        return

    for row_index in range(2, sheet.max_row + 1):
        item_code = sheet.cell(row=row_index, column=1).value
        if not item_code:
            continue


        prices = get_item_prices(item_code)

        for city, col_idx in city_columns.items():
            price_to_write = prices.get(city, '-')
            sheet.cell(row=row_index, column=col_idx).value = price_to_write

        time.sleep(0.1)
    print(f"--- ✅ Завершено оновлення аркуша: '{sheet.title}' ---")


# --- ГОЛОВНА ЧАСТИНА ---
if __name__ == "__main__":
    excel_file_name = "test.xlsx"

    try:
        workbook = openpyxl.load_workbook(excel_file_name)
        sheet_names = workbook.sheetnames
    except FileNotFoundError:
        print(f"❌ Помилка: Файл '{excel_file_name}' не знайдено.")
        exit()

    print("Доступні аркуші в файлі:")
    for i, name in enumerate(sheet_names):
        print(f"  {i + 1}: {name}")
    print("\n  0: Оновити ВСІ аркуші (паралельно)")

    try:
        choice = int(input("\nВведіть номер аркуша для оновлення: "))

        if choice == 0:
            # --- НОВА ЛОГІКА ДЛЯ ПАРАЛЕЛЬНОГО ОНОВЛЕННЯ ---
            # Ви можете змінити max_workers, щоб контролювати кількість одночасних потоків
            # Більше потоків = швидше, але більше навантаження на сайт
            max_workers = 5
            print(f"\n🚀 Запускаємо паралельне оновлення всіх аркушів у {max_workers} потоках...")

            with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as executor:
                # Створюємо задачу для кожного аркуша в книзі
                # workbook є ітератором, який повертає об'єкти аркушів
                executor.map(update_sheet, workbook)

        elif 1 <= choice <= len(sheet_names):
            # Оновлення одного вибраного аркуша (як і раніше)
            selected_sheet = workbook[sheet_names[choice - 1]]
            update_sheet(selected_sheet)
        else:
            print("❌ Неправильний вибір. Будь ласка, введіть число зі списку.")
            exit()

    except ValueError:
        print("❌ Помилка: потрібно ввести число.")
        exit()

    try:
        workbook.save(excel_file_name)
        print(f"\n💾 Всі оновлення завершено! Файл '{excel_file_name}' збережено.")
    except PermissionError:
        print(
            f"\n❌ Помилка збереження: Не вдалося зберегти файл '{excel_file_name}'. Можливо, він відкритий в Excel. Закрийте його та спробуйте знову.")